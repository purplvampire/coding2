import openpyxl
wb = openpyxl.Workbook()
x = wb.get_sheet_names()
print(x)
# rename sheet
sheet = wb.active
sheet.title = 'Sapm Bacon Eggs Sheet'
x = wb.get_sheet_names()
print(x)
# create sheet
wb.create_sheet(index=0, title='First Sheet')
wb.create_sheet(index=2, title='Middle Sheet')
x = wb.get_sheet_names()
print(x)
# remove sheet
wb.remove_sheet(wb.get_sheet_by_name('First Sheet'))
wb.remove_sheet(wb.get_sheet_by_name('Middle Sheet'))
x = wb.get_sheet_names()
print(x)
# insert data to cell
sheet = wb.get_sheet_by_name('Sapm Bacon Eggs Sheet')
sheet['A1'] = 'Hello World!'
y = sheet['A1'].value
print(y)
# Setup font
from openpyxl.styles import Font
fontStyle = Font(sz=24, i=True)
sheet['A2'] = 'Hello World!'
sheet['A2'].font = fontStyle
# 設定列高與欄寬
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['A'].width = 30
sheet.column_dimensions['B'].width = 0.5
# Merge Cell
sheet.merge_cells('A1:C1')
sheet['A1'] = 'Cells merged together.'
sheet.merge_cells('A2:C5')
sheet['A2'] = 'Cells merged together.'
# Unmerge Cell
sheet.unmerge_cells('A2:C5')
# Freeze Panes
sheet.freeze_panes = 'A2'
# save file
wb.save('example_copy.xlsx')


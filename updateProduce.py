#! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet.

import openpyxl
# Read Excel File
wb = openpyxl.load_workbook('produceSales.xlsx')
sheets = wb.get_sheet_names()
print(sheets)
# Indicate sheet
sheet = wb.active
print(sheet)
# Another word
sheet = wb.get_sheet_by_name('Sheet')
print(sheet)

# The produce types and their updated prices
PRICE_UPDATES = {'Garlic': '3.07',
                 'Celery': '1.19',
                 'Lemon': '1.27'}

# Loop through the rows and update the prices.
for rowNum in range(2, sheet.max_row):  # skip the first row
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

# Save to New Excel
wb.save('updateProduceSales.xlsx')





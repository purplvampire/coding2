#! python3
# sendDuesReminders.py - Sends emails based on payment status in spreadsheet.

import openpyxl, smtplib, sys

# Open the spreadsheet and get the latest dues status.
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.active
lastCol = sheet.max_column

latesMonth = sheet.cell(row=1, column=lastCol).value

# Check each member's payment status.
unpaidMembers = {}
for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

# Get SMTP password.
password = []
with open('mailPassword.txt', 'r') as f:
        for pw in f:
                password.append(pw)

# Log in to email account.
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('purplvampire@gmail.com', password[0])
print('Login Success!')
# Send out reminder emails.
for name, email in unpaidMembers.items():
        body = "Subject: %s dues unpaid.\nDear %s, Records show that you have not paid dues for %s. Please make this payment as soon as passible. Thank you!" % (latesMonth, name, latesMonth)
        print('Sending email to %s...' % email)
        sendmailStatus = smtpObj.sendmail('purplvampire@gmail.com', email, body)

        if sendmailStatus != {}:
                print('There was a problem sending email to %s: %s' % (email, sendmailStatus))
smtpObj.quit()
